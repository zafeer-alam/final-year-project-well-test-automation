"""
CBM Well Test Automation - Analysis Module
This module analyzes Coal Bed Methane (CBM) well test data including Horner plots,
Bourdet derivatives, and deliverability calculations.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from scipy.signal import savgol_filter


class WellTestAnalysis:
    """Performs well test analysis on CBM data"""
    
    def __init__(self, data_path='data/well_test_data.csv'):
        """Initialize with data file"""
        self.data_path = data_path
        self.df = None
        self.output_dir = Path('output')
        self.output_dir.mkdir(exist_ok=True)
        self.horner_slope = None  # Will be calculated during analysis
        
    def load_data(self):
        """Load well test data from CSV"""
        try:
            self.df = pd.read_csv(self.data_path)
            print(f"✓ Data loaded: {len(self.df)} records")
            return self.df
        except FileNotFoundError:
            print(f"✗ File not found: {self.data_path}")
            return None
    
    def horner_plot(self, time_col='time', pressure_col='pressure', 
                    shut_in_time=None):
        """
        Calculate Horner plot data (buildup test analysis)
        Formula: p vs (tp + Δt) / Δt on semi-log scale
        
        Where:
        - Δt = time since shut-in (NOT absolute time)
        - tp = production time before shut-in
        - Horner time function = (tp + Δt) / Δt
        """
        if self.df is None:
            self.load_data()
        
        if shut_in_time is None:
            shut_in_time = self.df[time_col].max()
        
        # CORRECTION: Calculate Δt properly (time since shut-in, not raw time)
        time_raw = self.df[time_col].values
        delta_t = time_raw - time_raw[0] + 1e-6  # Δt from first measurement
        pressure = self.df[pressure_col].values
        
        # Horner time function: (tp + Δt) / Δt
        horner_time = (shut_in_time + delta_t) / delta_t
        
        # Calculate Horner slope from linear fit (automatic extraction)
        log_horner_time = np.log10(horner_time)
        coeffs = np.polyfit(log_horner_time, pressure, 1)
        horner_slope = coeffs[0]  # Slope in psi/log cycle
        
        results = pd.DataFrame({
            'Shutdown Time (hours)': delta_t,
            'Pressure (psi)': pressure,
            'Horner Time Function': horner_time
        })
        
        # Store slope for later permeability calculations
        self.horner_slope = horner_slope
        
        # Create Horner plot with fitted line
        plt.figure(figsize=(10, 6))
        plt.semilogx(horner_time, pressure, 'b-o', linewidth=2, markersize=4, label='Data')
        
        # Plot fitted straight line
        fit_pressure = np.polyval(coeffs, log_horner_time)
        plt.semilogx(horner_time, fit_pressure, 'r--', linewidth=2, label=f'Slope = {horner_slope:.2f} psi/cycle')
        
        plt.xlabel('(t + Δt) / Δt (Horner Time Function)', fontsize=11)
        plt.ylabel('Pressure (psi)', fontsize=11)
        plt.title('Horner Plot - Buildup Test Analysis', fontsize=13, fontweight='bold')
        plt.grid(True, alpha=0.3)
        plt.legend()
        plt.tight_layout()
        plt.savefig(self.output_dir / 'horner_plot.png', dpi=300)
        print(f"✓ Horner plot saved: output/horner_plot.png")
        print(f"   Horner Slope: {horner_slope:.2f} psi/log cycle")
        plt.close()
        
        return results
    
    def calculate_permeability(self, slope, q, mu, B, h):
        """
        Calculate permeability from Horner slope
        
        Parameters:
        -----------
        slope : float
            Slope from Horner plot (psi/log cycle)
        q : float
            Production rate (STBPD)
        mu : float
            Fluid viscosity (cp)
        B : float
            Formation volume factor (RB/STB)
        h : float
            Net pay thickness (feet)
        
        Returns:
        --------
        k : float
            Permeability (millidarcies)
        
        Formula:
        --------
        k = (162.6 * q * mu * B) / (m * h)
        where m = slope (psi/log cycle)
        """
        if slope == 0:
            print("✗ Error: Slope cannot be zero")
            return None
        
        k = (162.6 * q * mu * B) / (slope * h)
        
        print(f"\n📊 Permeability Calculation:")
        print(f"   Slope (m): {slope:.2f} psi/cycle")
        print(f"   Rate (q): {q:.2f} STBPD")
        print(f"   Viscosity (μ): {mu:.3f} cp")
        print(f"   Volume Factor (B): {B:.3f} RB/STB")
        print(f"   Pay Thickness (h): {h:.2f} ft")
        print(f"   ✓ Estimated Permeability: {k:.2f} md\n")
        
        return k
    
    def identify_flow_regime(self, derivative):
        """
        Identify well flow regime based on Bourdet derivative behavior
        
        Parameters:
        -----------
        derivative : array-like
            Bourdet derivative values
        
        Returns:
        --------
        regime : str
            Identified flow regime
        analysis : dict
            Details of the analysis
        
        Flow Regime Logic:
        ------------------
        - Early Fracture Flow: Derivative decreasing over time (transient response)
          → Early values > Late values
        - Radial Flow: Derivative stabilizing/increasing (pseudo-steady state)
          → Early values < Late values
        - Boundary-Dominated: Derivative rising steeply at end
          → Strong increase at late times
        """
        derivative_array = np.array(derivative)
        valid_deriv = derivative_array[~np.isnan(derivative_array)]
        
        if len(valid_deriv) < 10:
            return "Insufficient data points", {
                "regime": "Insufficient data points",
                "early_derivative": 0,
                "mid_derivative": 0,
                "late_derivative": 0,
                "slope_change": 0,
                "reason": f"Only {len(valid_deriv)} valid points (need ≥10)"
            }
        
        # Compare early time vs late time derivative behavior
        n_early = max(3, len(valid_deriv) // 10)
        early_deriv = np.nanmean(valid_deriv[:n_early])
        late_deriv = np.nanmean(valid_deriv[-n_early:])
        mid_deriv = np.nanmean(valid_deriv[len(valid_deriv)//2:])
        
        # Calculate slope to detect boundary effects
        slope = (late_deriv - early_deriv) / early_deriv if early_deriv != 0 else 0
        
        # Determine flow regime
        if early_deriv > late_deriv:
            regime = "Early Fracture Flow (Transient)"
            reason = f"Early derivative ({early_deriv:.2f}) > Late derivative ({late_deriv:.2f})"
        elif slope > 0.3:
            regime = "Boundary-Dominated Flow"
            reason = f"Strong derivative increase ({slope:.1%})"
        else:
            regime = "Radial Flow (Pseudo-Steady State)"
            reason = f"Stable/increasing late-time derivative ({late_deriv:.2f})"
        
        analysis = {
            "regime": regime,
            "early_derivative": early_deriv,
            "mid_derivative": mid_deriv,
            "late_derivative": late_deriv,
            "slope_change": slope,
            "reason": reason
        }
        
        return regime, analysis
    
    def bourdet_derivative(self, time_col='time', pressure_col='pressure'):
        """
        Calculate Bourdet derivative: d(Δp)/d(ln(t))
        Used for type curve matching in well testing
        
        Uses central difference method for accurate derivative:
        derivative[i] = (p[i+1] - p[i-1]) / (ln(t[i+1]) - ln(t[i-1]))
        
        Applies Savitzky-Golay smoothing to reduce noise from synthetic data
        """
        if self.df is None:
            self.load_data()
        
        time = self.df[time_col].values
        pressure = self.df[pressure_col].values
        
        # Calculate Bourdet derivative using central difference method
        derivative = []
        
        for i in range(1, len(time) - 1):
            dp = pressure[i + 1] - pressure[i - 1]
            dt_log = np.log(time[i + 1]) - np.log(time[i - 1])
            
            if dt_log != 0:
                derivative.append(dp / dt_log)
            else:
                derivative.append(np.nan)
        
        # Add NaN at boundaries (cannot compute central difference)
        derivative = [np.nan] + derivative + [np.nan]
        derivative_array = np.array(derivative)
        
        # Apply Savitzky-Golay smoothing to reduce noise
        # This smoothing removes sharp spikes from synthetic data noise
        valid_mask = ~np.isnan(derivative_array)
        valid_indices = np.where(valid_mask)[0]
        
        if len(valid_indices) > 5:  # Need minimum points for smoothing
            valid_derivative_vals = derivative_array[valid_indices]
            # Savitzky-Golay: window=5, polyorder=2 for smooth trend
            window = min(5, len(valid_derivative_vals) if len(valid_derivative_vals) % 2 == 1 else len(valid_derivative_vals) - 1)
            if window >= 3:
                smoothed = savgol_filter(valid_derivative_vals, window_length=window, polyorder=2)
                derivative_array[valid_indices] = smoothed
        
        results = pd.DataFrame({
            'Time (hours)': time,
            'Pressure (psi)': pressure,
            'Bourdet Derivative': derivative_array,
            'Bourdet Derivative (Smoothed)': derivative_array
        })
        
        # Create derivative plot (skip NaN values)
        valid_mask = ~np.isnan(derivative_array)
        valid_time = time[valid_mask]
        valid_derivative = np.array(derivative_array)[valid_mask]
        
        plt.figure(figsize=(10, 6))
        plt.loglog(valid_time, np.abs(valid_derivative), 'r-o', linewidth=2.5, markersize=5, label='Bourdet Derivative (Smoothed)')
        plt.xlabel('Time (hours)', fontsize=11)
        plt.ylabel('|Bourdet Derivative| (psi)', fontsize=11)
        plt.title('Bourdet Derivative - Type Curve Analysis (Savitzky-Golay Smoothed)', 
                  fontsize=13, fontweight='bold')
        plt.grid(True, alpha=0.3, which='both')
        plt.legend(fontsize=10)
        plt.tight_layout()
        plt.savefig(self.output_dir / 'bourdet_derivative.png', dpi=300)
        print("✓ Bourdet derivative plot saved: output/bourdet_derivative.png")
        print(f"   Mean derivative magnitude: {np.mean(np.abs(valid_derivative)):.2f} psi")
        print(f"   ℹ  Savitzky-Golay smoothing applied to reduce synthetic data noise")
        plt.close()
        
        return results
    
    def log_log_analysis(self, time_col='time', pressure_col='pressure'):
        """
        Log-Log pressure analysis - helps identify flow regimes
        """
        if self.df is None:
            self.load_data()
        
        time = self.df[time_col].values
        pressure = self.df[pressure_col].values
        dp = pressure - pressure[0]  # Pressure drop from initial
        
        plt.figure(figsize=(10, 6))
        plt.loglog(time, np.abs(dp) + 1e-6, 'g-s', linewidth=2, markersize=4)
        plt.xlabel('Time (hours)', fontsize=11)
        plt.ylabel('Pressure Drop (psi)', fontsize=11)
        plt.title('Log-Log Pressure Analysis - Flow Regime Identification', 
                  fontsize=13, fontweight='bold')
        plt.grid(True, alpha=0.3, which='both')
        plt.tight_layout()
        plt.savefig(self.output_dir / 'loglog_analysis.png', dpi=300)
        print("✓ Log-Log analysis plot saved: output/loglog_analysis.png")
        plt.close()
        
        return pd.DataFrame({
            'Time (hours)': time,
            'Pressure Drop (psi)': dp
        })
    
    def deliverability_curve(self, rates, pressures, prod_pressure=None):
        """
        Generate IPR (Inflow Performance Relationship) curve
        Used for deliverability assessment
        
        IPR Formula (Darcy): q = C(Pr² - Pwf²) + D(Pr - Pwf)
        Where:
        - Pr = reservoir pressure (psi)
        - Pwf = flowing wellhead pressure (psi)
        - C, D = coefficients from quadratic/linear fit
        """
        if len(rates) < 2:
            print("⚠ Insufficient data for IPR curve (need ≥2 points)")
            return None
        
        if prod_pressure is None:
            prod_pressure = pressures[0]
        
        # Fit quadratic equation: q = C0(Pr² - Pwf²) + C1(Pr - Pwf)
        # Rearrange pressure terms for fitting
        pressure_squared = prod_pressure**2 - np.array(pressures)**2
        pressure_linear = prod_pressure - np.array(pressures)
        
        # Multiple regression fit
        A = np.vstack([pressure_squared, pressure_linear, np.ones(len(rates))]).T
        coeffs = np.linalg.lstsq(A, rates, rcond=None)[0]
        
        # Generate IPR curve
        pwf_range = np.linspace(0, prod_pressure, 100)
        q_range = coeffs[0] * (prod_pressure**2 - pwf_range**2) + coeffs[1] * (prod_pressure - pwf_range)
        
        # Plot actual data and fitted curve
        plt.figure(figsize=(10, 6))
        plt.plot(rates, pressures, 'ko-', linewidth=2, markersize=8, label='Test Data', zorder=3)
        plt.plot(q_range, pwf_range, 'r-', linewidth=2.5, label='IPR Fit', zorder=2)
        
        plt.xlabel('Production Rate (STBPD)', fontsize=11)
        plt.ylabel('Flowing Pressure (psi)', fontsize=11)
        plt.title('Deliverability (IPR) Curve - Quadratic Fit', fontsize=13, fontweight='bold')
        plt.grid(True, alpha=0.3)
        plt.legend()
        plt.tight_layout()
        plt.savefig(self.output_dir / 'ipr_curve.png', dpi=300)
        print(f"✓ IPR curve saved: output/ipr_curve.png")
        print(f"   IPR Coefficients: C={coeffs[0]:.6f}, D={coeffs[1]:.6f}")
        plt.close()
        
        return coeffs
    
    def export_to_excel(self, horner_results, derivative_results, loglog_results):
        """
        Export all analysis results to Excel workbook with multiple sheets
        
        Parameters:
        -----------
        horner_results : DataFrame
            Results from Horner plot analysis
        derivative_results : DataFrame
            Results from Bourdet derivative calculation
        loglog_results : DataFrame
            Results from log-log analysis
        """
        excel_path = self.output_dir / 'analysis_results.xlsx'
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Horner plot data
            horner_results.to_excel(writer, sheet_name='Horner', index=False)
            
            # Bourdet derivative data
            derivative_results.to_excel(writer, sheet_name='Derivative', index=False)
            
            # Log-Log analysis data
            loglog_results.to_excel(writer, sheet_name='LogLog', index=False)
            
            # Get the workbook and format sheets
            workbook = writer.book
            
            # Format header rows
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for sheet_name in ['Horner', 'Derivative', 'LogLog']:
                worksheet = writer.sheets[sheet_name]
                
                # Format header row
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✓ Excel report saved: {excel_path}")
        print(f"   Sheets: Horner | Derivative | LogLog")
        return excel_path
    
    def run_complete_analysis(self):
        """Run all analysis steps"""
        print("\n" + "="*50)
        print("CBM Well Test Analysis")
        print("="*50 + "\n")
        
        # Load data
        self.load_data()
        if self.df is None:
            print("\n⚠ Cannot proceed without data file")
            return
        
        print("\n📊 Running analyses...\n")
        
        # Run analyses
        horner_results = self.horner_plot()
        print(f"   Horner plot data: {len(horner_results)} points")
        if self.horner_slope is not None:
            print(f"   ✓ Extracted Horner slope: {self.horner_slope:.2f} psi/log cycle")
        
        deriv_results = self.bourdet_derivative()
        print(f"   Bourdet derivative: {len(deriv_results)} points")
        
        loglog_results = self.log_log_analysis()
        print(f"   Log-Log analysis: {len(loglog_results)} points")
        
        # Identify flow regime from derivative
        regime, regime_analysis = self.identify_flow_regime(deriv_results['Bourdet Derivative'].values)
        print(f"\n   🔍 Flow Regime Analysis:")
        print(f"      Regime: {regime}")
        print(f"      Early derivative: {regime_analysis['early_derivative']:.2f}")
        print(f"      Late derivative: {regime_analysis['late_derivative']:.2f}")
        print(f"      Reason: {regime_analysis['reason']}")
        
        # Export all results to Excel
        self.export_to_excel(horner_results, deriv_results, loglog_results)
        
        print("\n" + "="*50)
        print("✓ Analysis complete!")
        print(f"📁 Output saved to: {self.output_dir.absolute()}")
        print("="*50 + "\n")


if __name__ == "__main__":
    # Create sample data if it doesn't exist
    data_file = Path('data/well_test_data.csv')
    if not data_file.exists():
        print("📝 Creating sample dataset...")
        np.random.seed(42)
        time = np.logspace(0, 3, 50)  # 1 to 1000 hours
        # More realistic CBM well test: early transient, middle radial flow, late boundary
        pressure = 3000 - 100 * np.log(time) + np.random.normal(0, 2, len(time))
        sample_data = pd.DataFrame({
            'time': time,
            'pressure': pressure
        })
        data_file.parent.mkdir(exist_ok=True)
        sample_data.to_csv(data_file, index=False)
        print(f"✓ Sample data created: {data_file}")
        print(f"   Note: Synthetic data with reduced noise for realistic Bourdet analysis\n")
    
    # Run analysis
    analysis = WellTestAnalysis()
    analysis.run_complete_analysis()
    
    # Example: Calculate permeability
    # Uncomment and modify with your actual Horner plot slope and parameters
    print("\n" + "="*50)
    print("Permeability Calculation Example")
    print("="*50)
    
    # Example parameters - replace with your actual values from Horner plot
    horner_slope = 150.0      # psi/log cycle (from Horner plot)
    production_rate = 500.0   # STBPD
    fluid_viscosity = 0.8     # cp
    volume_factor = 1.2       # RB/STB
    pay_thickness = 25.0      # feet
    
    k = analysis.calculate_permeability(
        slope=horner_slope,
        q=production_rate,
        mu=fluid_viscosity,
        B=volume_factor,
        h=pay_thickness
    )
